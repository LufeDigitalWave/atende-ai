"""Generate WhatsApp Agents pricing calculator XLSX."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Sheet 1: Calculadora ---
ws = wb.active
ws.title = "Calculadora"

header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1E2761")
input_font = Font(bold=True, color="0000FF")
label_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

def style_input(cell):
    cell.font = input_font

# Title
ws["A1"] = "Calculadora de Pricing — WhatsApp Agents"
ws["A1"].font = Font(bold=True, size=16)
ws.merge_cells("A1:F1")

# Section: Inputs
ws["A3"] = "INPUTS DO PROJETO"
ws["A3"].font = label_font

labels = [
    ("A4", "Pacote", "B4", "Starter"),
    ("A5", "Conversas/mês estimadas", "B5", 500),
    ("A6", "Custo LLM por conversa (R$)", "B6", 0.04),
    ("A7", "Custo infra/mês (R$)", "B7", 150),
    ("A8", "Implantação cobrada (R$)", "B8", 2500),
    ("A9", "Mensalidade cobrada (R$)", "B9", 600),
    ("A10", "Horas de trabalho estimadas", "B10", 30),
    ("A11", "Valor hora desejado (R$/h)", "B11", 80),
]

for label_cell, label_text, value_cell, value in labels:
    ws[label_cell] = label_text
    ws[value_cell] = value
    style_input(ws[value_cell])

# Section: Calculated
ws["A13"] = "RESULTADOS"
ws["A13"].font = label_font

results = [
    ("A14", "Custo LLM mensal (R$)", "B14", "=B5*B6"),
    ("A15", "Custo total operação/mês (R$)", "B15", "=B14+B7"),
    ("A16", "Lucro mensal bruto (R$)", "B16", "=B9-B15"),
    ("A17", "Margem mensal (%)", "B17", "=IF(B9>0,B16/B9,0)"),
    ("A18", "Custo de entrega (R$)", "B18", "=B10*B11"),
    ("A19", "Lucro implantação (R$)", "B19", "=B8-B18"),
    ("A20", "Margem implantação (%)", "B20", "=IF(B8>0,B19/B8,0)"),
    ("A21", "Payback do cliente (meses de lucro mensal)", "B21", "=IF(B16>0,B8/B16,99)"),
    ("A22", "Revenue anual estimado (R$)", "B22", "=B8+(B9*12)"),
]

for label_cell, label_text, value_cell, formula in results:
    ws[label_cell] = label_text
    ws[value_cell] = formula

# Format percentages
for cell in ["B17", "B20"]:
    ws[cell].number_format = "0.0%"

# Format currency
for cell in ["B14", "B15", "B16", "B18", "B19", "B22"]:
    ws[cell].number_format = 'R$ #,##0'

ws["B21"].number_format = "0.0"

# Column widths
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 18

# --- Sheet 2: Pacotes ---
ws2 = wb.create_sheet("Pacotes")

headers = ["Pacote", "Implantação min", "Implantação max", "Mensalidade min", "Mensalidade max", "Prazo dias", "Perfil"]
for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    style_header(cell)

data = [
    ["Starter", 1500, 3000, 300, 800, "7-10", "1 agente, 1 fluxo, handoff"],
    ["Pro", 3500, 8000, 800, 2000, "10-20", "RAG, CRM/agenda, métricas"],
    ["Business", 8000, 20000, 2000, 5000, "20-40", "Multiagente, integrações, operação"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if isinstance(value, int) and col_idx in (2, 3, 4, 5):
            cell.number_format = 'R$ #,##0'

for col in range(1, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# --- Sheet 3: Custos LLM ---
ws3 = wb.create_sheet("Custos LLM")

headers3 = ["Modelo", "Input/1M tokens", "Output/1M tokens", "Custo por conversa (8 turnos)", "Obs"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    style_header(cell)

llm_data = [
    ["Claude Haiku 4.5", 0.80, 4.00, 0.02, "Melhor custo-benefício"],
    ["Claude Sonnet 4.6", 3.00, 15.00, 0.08, "Alta qualidade"],
    ["GPT-4o-mini", 0.15, 0.60, 0.01, "Mais barato"],
    ["GPT-4.1-mini", 0.40, 1.60, 0.02, "Factory/extractor"],
    ["GPT-4.1", 2.00, 8.00, 0.05, "Complexo"],
]

for row_idx, row_data in enumerate(llm_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=value)

for col in range(1, 6):
    ws3.column_dimensions[get_column_letter(col)].width = 22

# Save
output = "C:/tmp/atende-ai/docs/commercial/whatsapp-agents-pricing.xlsx"
wb.save(output)
print(f"Saved: {output}")
